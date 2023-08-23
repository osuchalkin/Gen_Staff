import openpyxl
import urllib.request


class Losses:
    """Получение xlsx-файла и данных из него"""

    def __init__(self):
        self.filename = 'genshtab.xlsx'
        self.url = "https://drive.usercontent.google.com/download?id=1kDJzRqv-Pmf3H3ziLuMyZeAOAglBg3zL&export=download&authuser=0&confirm=t&uuid=9c563a59-530b-429a-adce-b517789bfb73&at=APZUnTVFtwKcNIDn9l2Kxb5w1mXq:1691862972979"
        self.get_file(self.url)
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)
        self.sheet_list = self.wb.sheetnames  # 0- main, 1 - months, 2 - summary
        self.headers = self.get_headers()
        self.months = self.read_month_data(1)

    def get_file(self, dload_url):
        """Get file from Google Drive"""
        data = urllib.request.urlopen(dload_url).read()
        with open(self.filename, 'wb') as file:
            file.write(data)

    def get_current_data(self):
        sheet = self.wb[self.sheet_list[0]]
        datas = []

        for col in sheet.iter_cols(min_col=2, max_col=2, values_only=True):
            for cell in col:
                if str(type(cell)) == "<class 'datetime.datetime'>":
                    cell = str(cell)[8:10] + '.' + str(cell)[5:7] + '.' + str(cell)[:4]
                if cell is None:
                    break
                datas.append(cell)
        last = len(datas) - 1
        return datas[last]
    
    def get_headers(self):
        """get headers"""
        sheet = self.wb[self.sheet_list[2]]
        headers = []

        for row in sheet.iter_rows(max_row=1, values_only=True):
            for cell in row:
                headers.append(cell)

        return headers
    
    def read_headers_and_total_losses(self):
        """read the headers and total loses"""
        sheet = self.wb[self.sheet_list[2]]
        losses = []
        headers_and_total_losses = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            for cell in row:
                losses.append(cell)

        for i in range(len(self.headers)):
            headers_and_total_losses[self.headers[i]] = losses[i]

        return headers_and_total_losses
    
    def read_month_data(self, choice):
        datas = []
        sheet = self.wb[self.sheet_list[1]]

        for col in sheet.iter_cols(min_col=choice, max_col=choice, values_only=True):
            for cell in col:
                if str(type(cell)) == "<class 'datetime.datetime'>":
                    cell = str(cell)[5:7] + '.' + str(cell)[:4]
                if cell is None:
                    break
                datas.append(cell)
        datas.pop(0)
        
        return datas